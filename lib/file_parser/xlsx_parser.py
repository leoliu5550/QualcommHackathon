from pathlib import Path
import openpyxl
from lib.file_parser.base_parser import BaseParser, ParseResult

class XlsxParser(BaseParser):
    """XLSX 檔案解析器"""

    def parse(self, file_path: Path) -> ParseResult:
        try:
            workbook = openpyxl.load_workbook(file_path)
            text_content = []
            char_count = 0

            for sheet_name in workbook.sheetnames:
                sheet_header = f"工作表: {sheet_name}\n"
                text_content.append(sheet_header)
                char_count += len(sheet_header)

                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = ', '.join([str(cell) if cell is not None else '' for cell in row]) + '\n'
                    if char_count + len(row_text) > self.char_limit:
                        remaining = self.char_limit - char_count
                        text_content.append(row_text[:remaining])
                        char_count += remaining
                        break
                    text_content.append(row_text)
                    char_count += len(row_text)

                if char_count >= self.char_limit:
                    break

            content = ''.join(text_content)
            truncated_content, is_truncated = self._truncate_content(content)

            return ParseResult(
                success=True,
                content=truncated_content,
                file_type='xlsx',
                original_length=len(content),
                truncated=is_truncated,
                file_path=str(file_path)
            )

        except Exception as e:
            return ParseResult(
                success=False,
                error=f"無法解析XLSX檔案: {str(e)}",
                file_path=str(file_path)
            )
