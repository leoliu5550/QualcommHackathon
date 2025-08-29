"""Excel spreadsheet parser.

Extracts data from Excel files for analysis.
"""

from pathlib import Path
import openpyxl
from fileorg.parsers.base import BaseParser, ParseResult


class XlsxParser(BaseParser):
    """Parser for Excel spreadsheets.
    
    Reads cell data from all sheets.
    Requires openpyxl library.
    """

    def parse(self, file_path: Path) -> ParseResult:
        """Extract data from Excel file.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            ParseResult with spreadsheet data as text
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            text_content = []
            char_count = 0

            for sheet_name in workbook.sheetnames:
                sheet_header = f"Sheet: {sheet_name}\n"
                text_content.append(sheet_header)
                char_count += len(sheet_header)

                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = (
                        ", ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                    )
                    if char_count + len(row_text) > self.char_limit:
                        remaining = self.char_limit - char_count
                        text_content.append(row_text[:remaining])
                        char_count += remaining
                        break
                    text_content.append(row_text)
                    char_count += len(row_text)

                if char_count >= self.char_limit:
                    break

            content = "".join(text_content)
            truncated_content, is_truncated = self._truncate_content(content)

            return ParseResult(
                success=True,
                content=truncated_content,
                file_type="xlsx",
                original_length=len(content),
                truncated=is_truncated,
                file_path=str(file_path),
            )

        except Exception as e:
            return ParseResult(
                success=False, error=f"Failed to parse XLSX: {str(e)}", file_path=str(file_path)
            )
