"""
Excel (.xlsx) file parser for spreadsheet data extraction.

Handles Excel files with automatic reading of all sheets.
Extracts cell values as text for AI analysis, with truncation support.
"""

from pathlib import Path

from openpyxl import load_workbook

from fileorg.file_ops.ports.parser_ports import IParser, ParserOutput


class ExcelParser(IParser):
    """Parser for Excel .xlsx files with robust reading and truncation."""

    def _truncate_content(self, content: str, char_limit: int) -> str:
        """Truncate content if it exceeds character limit."""
        if len(content) > char_limit:
            return content[:char_limit]
        return content

    def parse(self, file_path: Path, char_limit: int) -> ParserOutput:
        """Extract text from Excel file (.xlsx) with truncation."""
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            content = ""

            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    # Convert all cells to string and join with commas
                    row_text = ",".join([str(cell) if cell is not None else "" for cell in row])
                    content += row_text + "\n"

                    # Stop if char_limit reached
                    if len(content) >= char_limit:
                        break
                if len(content) >= char_limit:
                    break

            truncated_content = self._truncate_content(content, char_limit)
            return ParserOutput(success=True, content=truncated_content, error="")

        except Exception as e:
            return ParserOutput(success=False, content="", error=f"Excel parsing failed: {e}")
